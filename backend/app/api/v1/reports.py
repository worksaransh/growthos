import csv
import io
from datetime import date, timedelta
from fastapi import APIRouter, Request
from fastapi.responses import StreamingResponse
from loguru import logger
from ...repositories.metrics_repo import get_metrics_summary, get_metrics
from ...repositories.product_repo import get_products
from ...repositories.customer_repo import get_customers
from ...core.database import fetch

router = APIRouter(prefix="/reports", tags=["reports"])


@router.get("/summary")
async def business_summary(
    request: Request,
    start_date: str | None = None,
    end_date: str | None = None,
):
    workspace_id = request.state.workspace_id
    today = date.today()
    sd = start_date or (today - timedelta(days=30)).isoformat()
    ed = end_date or today.isoformat()

    logger.info(f"Generating business summary for workspace {workspace_id} [{sd} to {ed}]")

    metrics = await get_metrics_summary(workspace_id, sd, ed)
    top_products = await get_products(workspace_id, sort_by="total_revenue", limit=10)
    top_customers = await get_customers(workspace_id, limit=10)

    # Ad spend by platform
    ad_breakdown = await fetch(
        """
        SELECT platform,
               SUM(spend) AS total_spend,
               SUM(impressions) AS impressions,
               SUM(clicks) AS clicks
        FROM ad_spend_daily
        WHERE workspace_id = $1
          AND spend_date BETWEEN $2 AND $3
        GROUP BY platform
        ORDER BY total_spend DESC
        """,
        workspace_id, sd, ed,
    )

    return {
        "period": {"start_date": sd, "end_date": ed},
        "metrics": metrics,
        "top_products": top_products,
        "top_customers": top_customers,
        "ad_breakdown": ad_breakdown,
    }


@router.get("/export/csv")
async def export_orders_csv(
    request: Request,
    start_date: str | None = None,
    end_date: str | None = None,
):
    workspace_id = request.state.workspace_id
    today = date.today()
    sd = start_date or (today - timedelta(days=30)).isoformat()
    ed = end_date or today.isoformat()

    logger.info(f"Exporting orders CSV for workspace {workspace_id} [{sd} to {ed}]")

    orders = await fetch(
        """
        SELECT shopify_order_id, shopify_order_number, order_status, financial_status,
               fulfillment_status, customer_name, customer_email, customer_id,
               gross_revenue, discount_amount, refund_amount, net_revenue,
               source_channel, payment_gateway, rto_status,
               tags, order_created_at, order_updated_at
        FROM shopify_orders
        WHERE workspace_id = $1
          AND DATE(order_created_at) BETWEEN $2 AND $3
        ORDER BY order_created_at DESC
        """,
        workspace_id, sd, ed,
    )

    output = io.StringIO()
    fieldnames = [
        "shopify_order_id", "shopify_order_number", "order_status", "financial_status",
        "fulfillment_status", "customer_name", "customer_email", "customer_id",
        "gross_revenue", "discount_amount", "refund_amount", "net_revenue",
        "source_channel", "payment_gateway", "rto_status",
        "tags", "order_created_at", "order_updated_at",
    ]
    writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for order in orders:
        row = dict(order)
        if isinstance(row.get("tags"), list):
            row["tags"] = ",".join(row["tags"])
        writer.writerow(row)

    output.seek(0)
    filename = f"orders_{sd}_to_{ed}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export/pdf")
async def export_pdf_report(
    request: Request,
    start_date: str | None = None,
    end_date: str | None = None,
):
    """Generate PDF business report."""
    workspace_id = request.state.workspace_id
    today = date.today()
    sd = start_date or (today - timedelta(days=30)).isoformat()
    ed = end_date or today.isoformat()

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.colors import HexColor
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        import io

        metrics = await get_metrics_summary(workspace_id, sd, ed)
        products = await get_products(workspace_id, sort_by="total_revenue", limit=10)

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4,
                                rightMargin=2*cm, leftMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)

        navy = HexColor("#0B1326")
        lavender = HexColor("#C0C1FF")
        gray = HexColor("#6B7280")

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("title", fontSize=24, textColor=lavender, spaceAfter=6, alignment=TA_LEFT, fontName="Helvetica-Bold")
        h2_style = ParagraphStyle("h2", fontSize=14, textColor=navy, spaceAfter=4, spaceBefore=12, fontName="Helvetica-Bold")
        body_style = ParagraphStyle("body", fontSize=10, textColor=gray, spaceAfter=4)

        elements = []
        elements.append(Paragraph("GrowthOS Business Report", title_style))
        elements.append(Paragraph(f"Period: {sd} to {ed}", body_style))
        elements.append(Spacer(1, 0.5*cm))

        # KPI Table
        kpi_data = [
            ["Metric", "Value"],
            ["Revenue", f"₹{metrics.get('revenue', 0):,.0f}"],
            ["Orders", str(metrics.get("orders", 0))],
            ["AOV", f"₹{metrics.get('aov', 0):,.0f}"],
            ["Customers", str(metrics.get("customers", 0))],
            ["Ad Spend", f"₹{metrics.get('ad_spend', 0):,.0f}"],
            ["ROAS", f"{metrics.get('roas', 0):.2f}x"],
        ]
        kpi_table = Table(kpi_data, colWidths=[8*cm, 8*cm])
        kpi_table.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), navy),
            ("TEXTCOLOR", (0,0), (-1,0), lavender),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,-1), 10),
            ("GRID", (0,0), (-1,-1), 0.5, gray),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [HexColor("#F8FAFC"), HexColor("#FFFFFF")]),
            ("PADDING", (0,0), (-1,-1), 8),
        ]))
        elements.append(Paragraph("Key Metrics", h2_style))
        elements.append(kpi_table)
        elements.append(Spacer(1, 0.5*cm))

        # Top Products
        if products:
            elements.append(Paragraph("Top Products by Revenue", h2_style))
            prod_data = [["Product", "Revenue", "Units"]]
            for p in products[:10]:
                prod_data.append([
                    str(p.get("title",""))[:40],
                    f"₹{p.get('total_revenue',0):,.0f}",
                    str(p.get("total_units",0))
                ])
            prod_table = Table(prod_data, colWidths=[10*cm, 5*cm, 3*cm])
            prod_table.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), navy),
                ("TEXTCOLOR", (0,0), (-1,0), lavender),
                ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE", (0,0), (-1,-1), 9),
                ("GRID", (0,0), (-1,-1), 0.5, gray),
                ("ROWBACKGROUNDS", (0,1), (-1,-1), [HexColor("#F8FAFC"), HexColor("#FFFFFF")]),
                ("PADDING", (0,0), (-1,-1), 6),
            ]))
            elements.append(prod_table)

        doc.build(elements)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=GrowthOS_Report_{sd}_{ed}.pdf"}
        )

    except ImportError:
        raise HTTPException(status_code=501, detail="PDF export requires: pip install reportlab")
    except Exception as e:
        logger.error(f"PDF export error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export/excel")
async def export_excel_report(
    request: Request,
    start_date: str | None = None,
    end_date: str | None = None,
):
    """Generate Excel business report with multiple sheets."""
    workspace_id = request.state.workspace_id
    today = date.today()
    sd = start_date or (today - timedelta(days=30)).isoformat()
    ed = end_date or today.isoformat()

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io

        metrics = await get_metrics_summary(workspace_id, sd, ed)
        products = await get_products(workspace_id, sort_by="total_revenue", limit=50)
        customers = await get_customers(workspace_id, limit=100)

        # Raw orders
        orders_raw = await fetch("""
            SELECT order_number, created_at, customer_email, customer_name,
                   total_price, total_discounts, financial_status, fulfillment_status,
                   utm_source, utm_medium, utm_campaign
            FROM shopify_orders
            WHERE workspace_id = $1
              AND created_at::date BETWEEN $2 AND $3
            ORDER BY created_at DESC
            LIMIT 10000
        """, workspace_id, sd, ed)

        wb = openpyxl.Workbook()

        navy_fill = PatternFill("solid", fgColor="0B1326")
        lavender_font = Font(color="C0C1FF", bold=True)
        header_font = Font(bold=True, color="0B1326")
        alt_fill = PatternFill("solid", fgColor="F8FAFC")
        thin = Side(style="thin", color="DDDDDD")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def style_header(ws, headers):
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.fill = navy_fill
                cell.font = lavender_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            ws.row_dimensions[1].height = 22

        def auto_width(ws):
            for col in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=10)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 40)

        # Sheet 1: Summary
        ws1 = wb.active
        ws1.title = "Summary"
        style_header(ws1, ["Metric", "Value", "Period"])
        summary_rows = [
            ("Revenue", f"₹{metrics.get('revenue',0):,.0f}", f"{sd} to {ed}"),
            ("Orders", metrics.get("orders",0), ""),
            ("AOV", f"₹{metrics.get('aov',0):,.0f}", ""),
            ("Unique Customers", metrics.get("customers",0), ""),
            ("Ad Spend", f"₹{metrics.get('ad_spend',0):,.0f}", ""),
            ("ROAS", f"{metrics.get('roas',0):.2f}x", ""),
            ("CAC", f"₹{metrics.get('cac',0):,.0f}", ""),
        ]
        for ri, row in enumerate(summary_rows, 2):
            for ci, val in enumerate(row, 1):
                cell = ws1.cell(row=ri, column=ci, value=val)
                cell.border = border
                if ri % 2 == 0:
                    cell.fill = alt_fill
        auto_width(ws1)

        # Sheet 2: Orders
        ws2 = wb.create_sheet("Orders")
        order_headers = ["Order #", "Date", "Customer", "Email", "Revenue", "Discount", "Status", "Fulfillment", "UTM Source", "UTM Medium", "UTM Campaign"]
        style_header(ws2, order_headers)
        for ri, o in enumerate(orders_raw, 2):
            vals = [
                o.get("order_number"), str(o.get("created_at",""))[:10],
                o.get("customer_name",""), o.get("customer_email",""),
                float(o.get("total_price",0)), float(o.get("total_discounts",0)),
                o.get("financial_status",""), o.get("fulfillment_status",""),
                o.get("utm_source",""), o.get("utm_medium",""), o.get("utm_campaign","")
            ]
            for ci, val in enumerate(vals, 1):
                cell = ws2.cell(row=ri, column=ci, value=val)
                cell.border = border
                if ri % 2 == 0:
                    cell.fill = alt_fill
        auto_width(ws2)

        # Sheet 3: Products
        ws3 = wb.create_sheet("Products")
        style_header(ws3, ["Product", "SKU", "Revenue", "Units Sold", "Avg Price", "Return Rate"])
        for ri, p in enumerate(products, 2):
            vals = [p.get("title",""), p.get("sku",""), float(p.get("total_revenue",0)),
                    int(p.get("total_units",0)), float(p.get("avg_price",0)), f"{p.get('return_rate',0):.1f}%"]
            for ci, val in enumerate(vals, 1):
                cell = ws3.cell(row=ri, column=ci, value=val)
                cell.border = border
                if ri % 2 == 0:
                    cell.fill = alt_fill
        auto_width(ws3)

        # Sheet 4: Customers
        ws4 = wb.create_sheet("Customers")
        style_header(ws4, ["Name", "Email", "Phone", "City", "Orders", "LTV", "Segment", "Last Order"])
        for ri, c in enumerate(customers, 2):
            vals = [c.get("name",""), c.get("email",""), c.get("phone",""),
                    c.get("city",""), int(c.get("order_count",0)),
                    float(c.get("total_spent",0)), c.get("rfm_segment",""),
                    str(c.get("last_order_at",""))[:10]]
            for ci, val in enumerate(vals, 1):
                cell = ws4.cell(row=ri, column=ci, value=val)
                cell.border = border
                if ri % 2 == 0:
                    cell.fill = alt_fill
        auto_width(ws4)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=GrowthOS_Report_{sd}_{ed}.xlsx"}
        )

    except ImportError:
        raise HTTPException(status_code=501, detail="Excel export requires: pip install openpyxl")
    except Exception as e:
        logger.error(f"Excel export error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/utm")
async def get_utm_attribution(
    request: Request,
    start_date: str | None = None,
    end_date: str | None = None,
):
    """Get UTM-based attribution summary."""
    workspace_id = request.state.workspace_id
    today = date.today()
    sd = start_date or (today - timedelta(days=30)).isoformat()
    ed = end_date or today.isoformat()

    rows = await fetch("""
        SELECT
            COALESCE(utm_source, 'direct') AS source,
            COALESCE(utm_medium, 'none') AS medium,
            COALESCE(utm_campaign, 'none') AS campaign,
            COUNT(*) AS orders,
            SUM(total_price) AS revenue,
            AVG(total_price) AS aov,
            COUNT(DISTINCT customer_email) AS unique_customers
        FROM shopify_orders
        WHERE workspace_id = $1
          AND created_at::date BETWEEN $2 AND $3
          AND financial_status NOT IN ('refunded', 'voided')
        GROUP BY utm_source, utm_medium, utm_campaign
        ORDER BY revenue DESC
        LIMIT 50
    """, workspace_id, sd, ed)
    return [dict(r) for r in rows]
